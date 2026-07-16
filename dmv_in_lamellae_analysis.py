"""
DMV in Lamellae Analysis

Analysis of DMV diameter, pore number, pore density,
and lamella geometry correction from cryo-ET data.

Authors:
- Montserrat Barcena
- Stanley Fronik

Parts of the analysis code were developed with the
assistance of ChatGPT (OpenAI) and subsequently reviewed,
tested, and modified by the authors.
"""

import math
from collections import defaultdict
import sys
import os
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter

# === Utility Functions ===

def euclidean_distance(p1, p2):
    return math.sqrt(sum((a - b) ** 2 for a, b in zip(p1, p2)))

def plane_from_points(points):
    points = np.array(points)
    centroid = points.mean(axis=0)
    _, _, vh = np.linalg.svd(points - centroid)
    normal = vh[-1]
    a, b, c = normal
    d = -np.dot(normal, centroid)
    return a, b, c, d

def plane_normal(plane):
    a, b, c, _ = plane
    norm = math.sqrt(a ** 2 + b ** 2 + c ** 2)
    return (a / norm, b / norm, c / norm)

def planes_are_parallel(plane1, plane2, tolerance=1e-3):
    n1 = plane_normal(plane1)
    n2 = plane_normal(plane2)
    dot_product = sum(a * b for a, b in zip(n1, n2))
    return abs(abs(dot_product) - 1) < tolerance

def distance_between_planes(plane1, plane2):
    if not planes_are_parallel(plane1, plane2):
        raise ValueError("Planes are not parallel")
    a, b, c, d1 = plane1
    _, _, _, d2 = plane2
    norm = math.sqrt(a ** 2 + b ** 2 + c ** 2)
    return abs(d1 - d2) / norm

def point_to_plane_distance(point, plane):
    a, b, c, d = plane
    x, y, z = point
    return abs(a * x + b * y + c * z + d) / math.sqrt(a ** 2 + b ** 2 + c ** 2)

def read_plane_points(plane_filename):
    upper, lower = [], []
    with open(plane_filename, 'r') as f:
        for line in f:
            if line.strip() and not line.strip().startswith('#'):
                parts = line.strip().split()
                plane_id = int(parts[0])
                coords = tuple(map(float, parts[2:5]))
                if plane_id == 1:
                    upper.append(coords)
                elif plane_id == 2:
                    lower.append(coords)
    return upper, lower

def spherical_cap_area(radius, height):
    return 2 * math.pi * radius * height

def sphere_surface_area(radius):
    return 4 * math.pi * radius ** 2

# === Logging redirection ===

class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for stream in self.streams:
            try:
                stream.write(data)
                stream.flush()
            except Exception:
                pass

    def flush(self):
        for stream in self.streams:
            try:
                stream.flush()
            except Exception:
                pass

# === Core Function ===

def process_single_file(vesicle_file, plane_file, nm_per_pixel, membrane_offset_px, log_stream):
    vesicle_data = defaultdict(list)
    with open(vesicle_file, 'r') as f:
        for line in f:
            if line.strip() and not line.strip().startswith('#'):
                parts = list(map(float, line.strip().split()))
                vesicle_id = int(parts[0])
                coords = tuple(parts[1:])
                vesicle_data[vesicle_id].append(coords)

    upper_points, lower_points = read_plane_points(plane_file)
    upper_plane = plane_from_points(upper_points)
    lower_plane = plane_from_points(lower_points)

    print(f"  Reading lamella points from: {plane_file}")
    print(f"  Upper plane coefficients: {upper_plane}")
    print(f"  Lower plane coefficients: {lower_plane}")

    thickness_px = distance_between_planes(upper_plane, lower_plane)
    thickness_nm = thickness_px * nm_per_pixel
    print(f"  Lamella thickness: {thickness_nm:.2f} nm")

    nm_per_pixel_squared = nm_per_pixel ** 2

    results = []

    for vesicle_id, coords in vesicle_data.items():
        print(f"\n  Vesicle {vesicle_id}:")

        if len(coords) < 2:
            print("    Not enough coordinates for vesicle.")
            continue

        center = coords[-1]
        pores = coords[:-1]
        num_pores = len(pores)

        print(f"    Center: {center}")
        distances = [euclidean_distance(center, pore) for pore in pores]
        for i, d in enumerate(distances, start=1):
            print(f"    Distance to pore {i}: {d:.2f} px")

        if distances:
            avg_radius = sum(distances) / len(distances)
            std_dev_radius = np.std(distances)
            print(f"    Diameter std dev (px): {2 * std_dev_radius:.4f}")
            diameter_px = 2 * avg_radius
            corrected_diameter_px = diameter_px + 2 * membrane_offset_px
            radius = corrected_diameter_px / 2
            diameter_nm = diameter_px * nm_per_pixel
            corrected_diameter_nm = corrected_diameter_px * nm_per_pixel
        else:
            avg_radius = diameter_px = corrected_diameter_px = radius = 0.0
            diameter_nm = corrected_diameter_nm = 0.0
            std_dev_radius = 0.0
            print(f"    Diameter std dev (px): 0.0000")

        dist_upper = point_to_plane_distance(center, upper_plane)
        dist_lower = point_to_plane_distance(center, lower_plane)
        print(f"    Distance to upper plane: {dist_upper:.2f} px")
        print(f"    Distance to lower plane: {dist_lower:.2f} px")

        surface_area_px2 = sphere_surface_area(radius)
        upper_cap_px2 = spherical_cap_area(radius, max(0, radius - dist_upper))
        lower_cap_px2 = spherical_cap_area(radius, max(0, radius - dist_lower))
        embedded_area_px2 = max(0.0, surface_area_px2 - upper_cap_px2 - lower_cap_px2)
        embedded_fraction = embedded_area_px2 / surface_area_px2 if surface_area_px2 > 0 else 0.0

        surface_area_nm2 = surface_area_px2 * nm_per_pixel_squared
        upper_cap_nm2 = upper_cap_px2 * nm_per_pixel_squared
        lower_cap_nm2 = lower_cap_px2 * nm_per_pixel_squared
        embedded_area_nm2 = embedded_area_px2 * nm_per_pixel_squared

        print(f"    Surface area (nm^2): {surface_area_nm2:.2f}")
        print(f"    Upper cap area (nm^2): {upper_cap_nm2:.2f}")
        print(f"    Lower cap area (nm^2): {lower_cap_nm2:.2f}")
        print(f"    Embedded surface area (nm^2): {embedded_area_nm2:.2f}")
        print(f"    Embedded fraction: {embedded_fraction:.4f}")

        estimated_pores = round(num_pores / embedded_fraction) if embedded_fraction > 0 else 0

        # Calculate density per 10000 nm^2, avoid div by zero
        if surface_area_nm2 > 0:
            density_per_10000nm2 = estimated_pores * 10000 / surface_area_nm2
        else:
            density_per_10000nm2 = 0.0

        print(f"    Density per 10000 nm^2: {density_per_10000nm2:.4f}")

        results.append((
            vesicle_id, diameter_px, diameter_nm,
            corrected_diameter_nm, 2 * std_dev_radius,
            num_pores, embedded_fraction, estimated_pores,
            surface_area_nm2, upper_cap_nm2, lower_cap_nm2, embedded_area_nm2,
            thickness_nm, density_per_10000nm2))

    return results

def process_file_list(list_filename, output_filename):
    import openpyxl

    with open(list_filename, 'r') as f:
        lines = [line.strip() for line in f if line.strip() and not line.startswith('#')]

    header = lines[0].split()
    nm_per_pixel = float(header[0])
    membrane_offset_nm = float(header[1])
    membrane_offset_px = membrane_offset_nm / nm_per_pixel
    input_files = [line.split() for line in lines[1:]]

    log_file = output_filename.replace('.txt', '.log')
    excel_file = output_filename.replace('.txt', '.xlsx')

    with open(output_filename, 'w') as out, open(log_file, 'w') as log_stream:
        sys.stdout = Tee(sys.__stdout__, log_stream)
        print(f"\nLogging session started. Log will be saved to '{log_file}'\n")
        print(f"Read nm_per_pixel: {nm_per_pixel}, membrane_offset_px: {membrane_offset_px}")
        print(f"Files to process: {input_files}")

        headers = [
            "File_Index", "Rootname", "Lamella_Thickness_nm", "Vesicle_ID", "Diameter_px", "Diameter_nm",
            "Corrected_Diameter_nm", "Diameter_StdDev_px", "Number_of_Pores",
            "Embedded_Fraction", "Estimated_Total_Pores",
            "Surface_nm2", "UpperCap_nm2", "LowerCap_nm2", "Embedded_nm2",
            "Density_per_10000nm2"
        ]

        fmt = ("{:<11}{:<15}{:<22.2f}{:<12}{:<13.2f}{:<13.2f}{:<23.2f}{:<20.4f}{:<18}"
               "{:<18.4f}{:<23}{:<15.2f}{:<15.2f}{:<15.2f}{:<15.2f}{:<23.4f}\n")

        out.write("".join(f"{h:<22}" for h in headers) + "\n")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vesicle Analysis"
        ws.append(headers)

        for i, (vesicle_file, plane_file) in enumerate(input_files, start=1):
            print(f"\nProcessing file {i}: {vesicle_file}")
            rootname = os.path.basename(vesicle_file).replace("-pores.txt", "")
            results = process_single_file(vesicle_file, plane_file, nm_per_pixel, membrane_offset_px, log_stream)
            for res in results:
                thickness_nm = res[-2]  # second last in tuple
                density = res[-1]       # last in tuple
                out.write(fmt.format(i, rootname, thickness_nm, *res[:-2], density))
                ws.append([i, rootname, thickness_nm] + list(res[:-2]) + [density])

        # Dynamically set Excel column widths based on header length
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(header) + 2  # padding
            ws.column_dimensions[col_letter].width = max_length

        wb.save(excel_file)
        print("\nLogging session finished.")

# === Main Program ===

if __name__ == "__main__":
    list_file = input("Enter the input list filename: ")
    output_file = input("Enter the output filename: ")
    process_file_list(list_file, output_file)
